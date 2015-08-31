#!/usr/bin/env python
# -*- coding: UTF-8 -*-
# coding=utf-8

import sys, re, os, sh, smtplib
from openpyxl import load_workbook
from datetime import datetime, timedelta
import gmail_client
import locale
import logging
from datetime import *
from dateutil.relativedelta import *
import json
import StringIO

# logging.basicConfig(level=logging.INFO)

class Tee(object):
    def __init__(self):
        self.file = StringIO.StringIO()
        self.stdout = sys.stdout
        sys.stdout = self
    def __del__(self):
        sys.stdout = self.stdout
        self.file.close()
    def write(self, data):
        self.file.write(data+"\n")
        self.stdout.write(data+"\n")

def Help(msg=""):
	print "Converts ahnlab Menu to Outlook compativle csv"
	if msg !="":
		print "Error: %s" % (msg)
	sys.exit(-1)

def send_email(title, body):
	FROM='alarm@1kko.com'
	TO=['me@1kko.com']
	SUBJECT=title
	TEXT=body

	with open('./config.json') as fd:
		logininfo=json.load(fd)

	GMAIL_ID=logininfo['GMAIL_ID']
	GMAIL_PW=logininfo['GMAIL_PW']

	message = """\From: %s\nTo: %s\nSubject: %s\n\n%s""" % (FROM, ", ".join(TO), SUBJECT, TEXT)

	try:
		#server = smtplib.SMTP(SERVER) 
		server = smtplib.SMTP("smtp.gmail.com", 587) #or port 465 doesn't seem to work!
		server.ehlo()
		server.starttls()
		server.login(GMAIL_ID, GMAIL_PW)
		server.sendmail(FROM, TO, message)
		#server.quit()
		server.close()
		print 'successfully sent the mail'
	except:
		print "failed to send mail"


def fetch_attachments():
	retval=[]
	g=gmail_client.Gmail()
	
	locale.setlocale(locale.LC_TIME,'C')
	
	TODAY=date.today()
	Before=TODAY+relativedelta(days=-5)
	After=TODAY

	with open('./config.json') as fd:
		logininfo=json.load(fd)

	GMAIL_ID=logininfo['GMAIL_ID']
	GMAIL_PW=logininfo['GMAIL_PW']

	g.login(GMAIL_ID,GMAIL_PW)
	#emails=g.inbox().mail(prefetch=True, after=datetime(2015,2,5), before=datetime(2015,2,20), sender='mk2926.jeong@ourhome.co.kr')
	emails=g.label('Ahnapp').mail(prefetch=True, after=After, before=Before, unread=True)

	for email in emails:
		print "New email found, checking attachments"
		for attachment in email.attachments:
			attname=attachment.name.encode('utf-8').strip()
			attsize=str(attachment.size).encode('utf-8').strip()

			filename, file_ext=os.path.splitext(attname)
			if file_ext==".xlsx":
				print 'saving attachment: %s (%s Bytes)' % (attname, attsize)
				attachment.save('./'+attname)
				retval.append(attname)
			else:
				print 'skip: %s' % attname

		email.mark_read()
		email.archive()

	g.logout()

	return retval


def range_finder(ws, Column, search_str, padding=0, heading=0, start=1, end=100):
	cell_start=None
	cell_end=None

	for i in range(start,end):
		cell=Column+str(i)

		if cell_start!=None:
			if ws[cell].value!=None :
				cell_end=i-1+padding;
				break;

		if cell_start==None:
			if ws[cell].value!=None:
				if search_str=="":
					cell_start=i+heading;
				elif ws[cell].value.find(search_str) > 0:
					cell_start=i+heading;

	retval={'column':Column, \
			'data':ws[Column+str(cell_start)].value, \
			'row':{ 'start':cell_start, 'end':cell_end}, \
			'cell':{ 'start':Column+str(cell_start), 'end':Column+str(cell_end)} \
			}

	return retval


def data_finder(ws, Column, start, end):
	retval=[]

	# print "data_finder: start=",start," end=",end
	# find start point first, then add end points.
	temps=[]
	for i in range(start, end*2):
		cell=Column+str(i)
		if ws[cell].value!=None:
			temps.append(i)

	#print "temps: ",temps
	j=1
	for i in range(start, end+1):
		cell=Column+str(i)
		if ws[cell].value!=None:
			if temps[j]-1 >= end:
				real_end=end
			else:
				real_end=temps[j]-1
			#real_end=end

			subval={'column':Column, \
					'data':ws[cell].value, \
					'row':{'start':i, 'end':real_end}, \
					'cell':{'start':cell, 'end':Column+str(real_end)} \
					}
			#print "j=",j,"end=",end, "temps[j]=",temps[j],"subval=",subval
			retval.append(subval)
			j+=1

	return retval


def data_merger(ws, Column, start, end, header=""):
	""" From given range of cells, return merged data.
	This does not actually merge cells in worksheet.
	"""
	# print "rx=", start, end
	retval=""
	header="\n["+header.replace("\n"," ")+"]"

	for i in range(start, end+1):
		cell=Column+str(i)
		# print "i=",cell
		if ws[cell].value==None:
			retval+="\n"
		else:
			retval+="\n"+ws[cell].value
	
	if len(retval)>1:
		return header+retval
	else:
		return ""


def convert(filename):
	tee.write("loading: %s " % filename)
	wb=load_workbook(filename)
	# activate worksheet
	ws=wb.active


	# # ----- 플러스메뉴 fix 시작
	plusMenuCells=[]
	for plusMenuItem in ws.merged_cell_ranges:
		cell=plusMenuItem.split(":")[0]

		if ws[cell].value!=None:
			if ws[cell].value.find(u'플러스')>=0:
				plusMenuCells.append(plusMenuItem)

	for time in ['07:30','11:30','17:30']:
		meal=range_finder(ws, "A", time)
		mealStart=meal['row']['start']
		mealEnd=meal['row']['end']

		for i in range(0,len(plusMenuCells)):
			# print "plusMenuCells:",plusMenuCells
			plusStart=int(plusMenuCells[i].split(":")[0].replace("A",""))
			plusEnd=int(plusMenuCells[i].split(":")[1].replace("B",""))

			# print "mealEnd: ", mealEnd, "  plusStart: ",plusStart
			if mealEnd == int(plusStart)-1:
				plusStart=str(plusStart)
				plusEnd=str(plusEnd)
				mealEnd=str(mealEnd)
				mealStart=str(mealStart)
				
				# print "unmerge: ",'A'+plusStart+':B'+plusEnd
				ws.unmerge_cells('A'+plusStart+':B'+plusEnd)
				
				# print "copy: ",'A'+plusStart,"->",'B'+plusStart, ws['A'+plusStart]
				ws['B'+plusStart].value=ws['A'+plusStart].value
				
				# print "umnerge: ",'A'+mealStart+':A'+mealEnd
				ws.unmerge_cells('A'+mealStart+':A'+mealEnd)
				
				# print "merge: ",'A'+mealStart+':A'+plusEnd
				ws.merge_cells('A'+mealStart+':A'+plusEnd)
				if plusEnd!=plusStart:
					ws.merge_cells('B'+plusStart+':B'+plusEnd)


	# # BreakfastMenu Fix
	# # 플러스 메뉴 병합 해제
	# ws.unmerge_cells('A14:B14')
	# # '플러스메뉴' (A14) 를 (B14)로 복사
	# ws['B14']=ws['A14'].value
	# # A14 클리어
	# ws['A14']=""
	# # 원래의 아침메뉴 Range를 병합해제
	# ws.unmerge_cells('A3:A13')
	# # 플러스 메뉴가 있는 부분까지 아침메뉴로 병합
	# ws.merge_cells('A3:A14')

	# ----- 플러스메뉴 fix 끝 

	# ----- Start of Last Row Data Fix.
	# 마지막에 글자를 넣어줘야 range를 구할 수 있음
	last_row=0
	for mcell in ws.merged_cells:
		cell=re.findall(r"[^\W\d_]+|\d+",mcell)
		if cell[0]=='A' and int(cell[1])>=last_row:
			last_row=int(cell[1])

	last_col='A'
	for mcell in ws.merged_cells:
		cell=re.findall(r"[^\W\d_]+|\d+",mcell)
		if ord(cell[0])>=ord(last_col):
			last_col=cell[0]


	last_row=str(last_row)
	# print "last_row:", last_row, "last_col:", last_col
	ws.unmerge_cells('A'+last_row+':'+last_col+last_row)

	end=ord(last_col)-ord('A')
	for i in range (0,end+1):
		col=chr(ord('A')+i)
		# print col+last_row
		ws[col+last_row]="."
	
	# ----- End of Last ROw Data Fix.

	# wb.save('balances.xlsx')

	# d=datetime.date.today()
	# year=d.year
	# month=ws['D2'].value.split(" ")[0].replace("월","")
	# day=ws['D2'].value.split(" ")[1].replace("일","")
	date_start=datetime.strptime( str(datetime.now().year)+' '+ ws['D2'].value.split(" ")[0].replace(u"\uc6d4","")+' '+ str(int(ws['D2'].value.split(" ")[1].replace(u"\uc77c",""))), '%Y %m %d')-timedelta(-1)
	
	filename_date_start=date_start+timedelta(-1)
	filename_date_end=filename_date_start+timedelta(6)

	outputfilename=str(filename_date_start.strftime('%Y%m%d'))+"-"+str(filename_date_end.strftime('%Y%m%d'))+".csv"
	outputfile=open(outputfilename, "w")

	tee.write("outputfilename: %s " % outputfilename)
	tee.write("outputURL: %s" % "http://926.1kko.com/menu/"+outputfilename)
	csvHeader=u"제목,시작 날짜,시작 시간,끝 날짜,끝 시간,하루 종일,미리 알림 설정/해제,미리 알림 날짜,미리 알림 시간,모임 이끌이,필수 참석자,선택 참석자,모임 리소스,거리,범주 항목,비용 정보,설명,숨김,시간 상태 보기,우선 순위,우편물 종류,장소"
	outputfile.write(csvHeader.encode('utf8')+"\n")


	# date_start.strftime('%Y-%m-%d')


	# To see what happened, uncomment follow
	# wb.save("converted.xlsx")

	# calculate range between breakfast, lunch, dinner
	numday=0
	#for col in ['C']:
	for col in ['C','D','E','F','G']:

		today=date_start+timedelta(numday)
		#for time in ['07:30']:
		for time in ['07:30','11:30','17:30']:

			if time=='07:30':
				ettd=60
				title=u"아침메뉴"
			if time=='11:30':
				ettd=90
				title=u"점심메뉴"
			if time=='17:30':
				ettd=90
				title=u"저녁메뉴"

			start_day=str(today.strftime('%m/%d/%Y'))
			start_time=time+":00"
			end_day=str(today.strftime('%m/%d/%Y'))
			end_time=datetime.strptime(time+":00",'%H:%M:%S')+timedelta(minutes=ettd)
			end_time=str(end_time.strftime('%H:%M:%S'))
			all_day="FALSE"
			notification_enabled="FALSE"
			notification_day=str(today.strftime('%m/%d/%Y'))
			notification_time=time
			meeting_host="Nine2Six"
			other_fields01=',,,,,,'
			other_fields02=u'FALSE,2,중간,보통,'

			#print "Date: ", today.strftime('%Y-%m-%d'), "Time: ", time, end_time.strftime('%H:%M')

			meal=range_finder(ws, "A", time)
			# print meal
			meal_category=data_finder(ws, "B", start=meal['row']['start'], end=meal['row']['end'])
			# print meal_category

			old_prnstr=""
			description=""
			for i in range(0,len(meal_category)):
				# print "meal_category[i]['data']", meal_category[i]['data'].strip("\n")
				# print meal_category[i]['row']['start'], meal_category[i]['row']['end']
				description+=data_merger(ws, col, meal_category[i]['row']['start'], meal_category[i]['row']['end'], header=meal_category[i]['data'])+"\n"
				
			new_prnstr=title+','+start_day+','+start_time+','+end_day+','+end_time+','+all_day+','+ \
				notification_enabled+','+notification_day+','+notification_time+','+meeting_host+','+ \
				other_fields01+'"'+description+'",'+other_fields02
			
			outputfile.write(new_prnstr.encode('utf8')+"\n")

			# skip on same line
			if old_prnstr!=new_prnstr :
				#print title+','+start_day+','+start_time+','+end_day+','+end_time+','+all_day+','+ \
				notification_enabled+','+notification_day+','+notification_time+','+meeting_host+','+ \
				other_fields01+'"'+description+'",'+other_fields02

				old_prnstr=new_prnstr

		numday+=1

	outputfile.close()
	return outputfilename


if __name__ == '__main__':
	# stdout = StringIO.StringIO()
	# sys.stdout = stdout
	if len(sys.argv)>=2:
		filename=sys.argv[1]
	else:
		sendEmailFlag=True
		try:
			commitFlag=False
			tee=Tee()

			git=sh.git.bake(_cwd='.')
			# print git.status()
			#### git stash, pull th
			tee.write("Check for update.")
			is_updated=git.pull()
			if (is_updated==0):
				pass
			# git.stash("remove")
			
			tee.write("Fetching from Email")
			attachments=set(fetch_attachments())
			for filename in attachments:
				try:
					convertedFilename=convert(filename)
					tee.write("New file found. Adding to repository: %s" % convertedFilename)
					git.add(convertedFilename)
					os.remove(filename)
					commitFlag=True
				except:
					tee.write("File is already removed or not exists")
					raise
			
			if commitFlag==True:
				tee.write ("Commiting changes")
				git.commit(m='menu_update')
				tee.write("Push to server")
				git.push()
				
				tee.write(str(git.status()))

				title="[Ahnapp] Menu Updated Successfully"
				body=tee.file.getvalue()
				# send_email(title,body)
			else:
				tee.write("nothing updated")
				# print tee.file.getvalue()
				sendEmailFlag=False
		except Exception as e:
			title="[Ahnapp] Error: Something went wrong ;("
			#body=tee.file.getvalue()+"\n"+str(e)
			body=str(e)
			raise

		finally:
			if sendEmailFlag==True: 
				send_email(title,body)


